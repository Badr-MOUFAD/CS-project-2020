from DataAndInstantiation import Instanciation
from openpyxl import workbook
from openpyxl import load_workbook

import datetime

Jorf = Instanciation.Jorf

# variable :
conteur = 0
engraisMp34_ = []
engraisJf_ = []
engraisJfExtraction = ["JFD", "JFF", "JFO", "JFQ", "JFT"]
commandes = []
CommandesEngraisMp34_ = []
conteur1 = 0
commandesJf = []
CommandesEngraisJf_ = []
conteur2 = 0
productionSulfurique = []
productionPhosphorique29 = []
productionPhosphorique54 = []
conteur3 = 0
productionSulfuriqueJf = []
productionPhosphorique29Jf = []
productionPhosphorique54Jf=[]
conteur4=0
capaciteEntite=[]
capaciteCommune=[]
AllEntite=["MP34","JFD","JFF","JFO","JFQ","JFT","PMP","IMACID"]
AllSouffre=[]
SouffreMp34=[]
SouffreJfd=[]
SouffreJff=[]
SouffreJfo=[]
SouffreJfq=[]
SouffreJft=[]
SouffrePmp=[]
SouffreImacid=[]
AllEntiteMp34=["JFD","JFF","JFO","JFQ","JFT","PMP","IMACID"]
Mp34from_=[]
Mp34FromJfd=[]
Mp34FromJff=[]
Mp34FromJfo=[]
Mp34FromJfq=[]
Mp34FromJft=[]
Mp34FromPmp=[]
Mp34FromImacid=[]
AllEntiteJfd=["JFF","JFO","JFQ","JFT","PMP","IMACID","MP34"]
Jfdfrom_=[]
JfdFromJff=[]
JfdFromJfo=[]
JfdFromJfq=[]
JfdFromJft=[]
JfdFromPmp=[]
JfdFromImacid=[]
JfdFromMp34=[]
AllEntiteJff=["MP34","JFD","JFO","JFQ","JFT","PMP","IMACID"]
jffFrom_=[]
jffFromImacid=[]
jffFromPmp=[]
jffFromJft=[]
jffFromJfq=[]
jffFromJfo=[]
jffFromMp34=[]
jffFromJfd=[]
AllEntiteJfo=["MP34","JFD","JFF","JFQ","JFT","PMP","IMACID"]
jfoFrom_=[]
jfoFromImacid=[]
jfoFromPmp=[]
jfoFromJft=[]
jfoFromJfq=[]
jfoFromJff=[]
jfoFromMp34=[]
jfoFromJfd=[]
AllEntiteJfq=["MP34","JFD","JFF","JFO","JFT","PMP","IMACID"]
jfqFrom_=[]
jfqFromImacid=[]
jfqFromPmp=[]
jfqFromJft=[]
jfqFromJfo=[]
jfqFromJff=[]
jfqFromMp34=[]
jfqFromJfd=[]
AllEntiteJft=["MP34","JFD","JFF","JFO","JFQ","PMP","IMACID"]
jftFrom_=[]
jftFromImacid=[]
jftFromPmp=[]
jftFromJfq=[]
jftFromJfo=[]
jftFromJff=[]
jftFromMp34=[]
jftFromJfd=[]
AllEntitePmp=["MP34","JFD","JFF","JFO","JFQ","JFT","IMACID"]
PmpFrom_=[]
PmpFromImacid=[]
PmpFromJft=[]
PmpFromJfq=[]
PmpFromJfo=[]
PmpFromJff=[]
PmpFromMp34=[]
PmpFromJfd=[]
AllEntiteImacid=["MP34","JFD","JFF","JFO","JFQ","JFT","PMP"]
imacidFrom_=[]
imacidFromPmp=[]
imacidFromJft=[]
imacidFromJfq=[]
imacidFromJfo=[]
imacidFromJff=[]
imacidFromMp34=[]
imacidFromJfd=[]

maintenanceD0 = []
listeRegime = []


#wb = load_workbook("C:/Users/HP/Desktop/input.xlsm",keep_vba=True)
wb = load_workbook("input.xlsm",keep_vba=True)
Sheet1 = wb["Paramétrage"]


# fonction :
def infoEngraisMp34(engraisMp34_):
    for ligneEngrais in Jorf['MP34'].lignesEngrais.values():
        r = ligneEngrais.régimeDeMarche
        d = ligneEngrais.disponibilité
        engraisMp34_.append([r[0], d])
    return engraisMp34_


def infoEngraisJf(engraisJf_):
    for i in engraisJfExtraction:
        for ligneEngrais in Jorf[i].lignesEngrais.values():
            r = ligneEngrais.régimeDeMarche
            d = ligneEngrais.disponibilité
            engraisJf_.append([r[0], d])
    return (engraisJf_)


def carnetMp34(commandes, CommandesEngraisMp34_, conteur1):
    for ligneEngrais in Jorf['MP34'].lignesEngrais.values():
        commandes.append(ligneEngrais.getCommandes())
    for i in commandes:
        conteur1 = conteur1 + 1

        for j in i:
            j.append(conteur1)
            CommandesEngraisMp34_.append(j)
    return CommandesEngraisMp34_


def carnetJf(commandesJf, conteur2, CommandesEngraisJf_):
    for e in engraisJfExtraction:
        for ligneEngrais in Jorf[e].lignesEngrais.values():
            commandesJf.append(ligneEngrais.getCommandes())
    for i in commandesJf:
        conteur2 = conteur2 + 1
        for j in i:
            j.append(conteur2)
            CommandesEngraisJf_.append(j)
    return CommandesEngraisJf_


def productionMp34(productionSulfurique, productionPhosphorique29, productionPhosphorique54, conteur3):
    for atelier in Jorf['MP34'].ateliers.values():
        for ligne in atelier.lignesDeProduction.values():

            if conteur3 < 6:
                productionSulfurique.append([ligne.régimeDeMarche[0], ligne.disponibilité])
            if conteur3 >= 6 and conteur3 < 12:
                productionPhosphorique29.append([ligne.régimeDeMarche[0], ligne.disponibilité])
            if conteur3 >= 12 and conteur3 < 32:
                productionPhosphorique54.append([ligne.régimeDeMarche[0], ligne.disponibilité])
            conteur3 = conteur3 + 1
    return productionSulfurique, productionPhosphorique29, productionPhosphorique54


def productionJf(productionSulfuriqueJf, productionPhosphorique29Jf, productionPhosphorique54Jf, conteur4):
    for nomAtelier in Jorf.ENCHAINEMENT_ATELIER:
        for entité in Jorf.entités.values():
            if entité.nom == 'MP34':
                continue
            atelier = entité.getAtelier(nomAtelier)
            for ligne in atelier.lignesDeProduction.values():
                if conteur4 < 7:
                    productionSulfuriqueJf.append([ligne.régimeDeMarche[0], ligne.disponibilité])
                if conteur4 >= 7 and conteur4 < 14:
                    productionPhosphorique29Jf.append([ligne.régimeDeMarche[0], ligne.disponibilité])
                if conteur4 >= 14 and conteur4 < 35:
                    productionPhosphorique54Jf.append([ligne.régimeDeMarche[0], ligne.disponibilité])
                conteur4 = conteur4 + 1
    return (productionSulfuriqueJf, productionPhosphorique29Jf, productionPhosphorique54Jf)


def capaciteUtilise(capaciteEntite, capaciteCommune):
    for entité in Jorf.entités.values():
        for stock in entité.stocks.values():
            capaciteEntite.append(stock.capacitéUtilisé[0])
    for stock in Jorf.stocksCommun.values():
        capaciteCommune.append(stock.capacitéUtilisé[0])
    return (capaciteEntite, capaciteCommune)


def echangeSouffre(AllSouffre, SouffreMp34, SouffreJfd, SouffreJff, SouffreJfo, SouffreJfq, SouffreJft, SouffrePmp,
                   SouffreImacid):
    for i in AllEntite:
        AllSouffre.append(Jorf[i].getEchanges('extérieur')[0])
    SouffreMp34.append(AllSouffre[0])
    SouffreJfd.append(AllSouffre[1])
    SouffreJff.append(AllSouffre[2])
    SouffreJfo.append(AllSouffre[3])
    SouffreJfq.append(AllSouffre[4])
    SouffreJft.append(AllSouffre[5])
    SouffrePmp.append(AllSouffre[6])
    SouffreImacid.append(AllSouffre[7])
    return (SouffreMp34, SouffreJfd, SouffreJff, SouffreJfo, SouffreJfq, SouffreJft, SouffrePmp, SouffreImacid)


def Mp34From(Mp34from_, Mp34FromJfd, Mp34FromJff, Mp34FromJfo, Mp34FromJfq, Mp34FromJft, Mp34FromPmp, Mp34FromImacid):
    for i in AllEntiteMp34:
        Mp34from_.append(Jorf["MP34"].getEchanges(i))
    Mp34FromJfd.append(Mp34from_[0])
    Mp34FromJff.append(Mp34from_[1])
    Mp34FromJfo.append(Mp34from_[2])
    Mp34FromJfq.append(Mp34from_[3])
    Mp34FromJft.append(Mp34from_[4])
    Mp34FromPmp.append(Mp34from_[5])
    Mp34FromImacid.append(Mp34from_[6])
    return (Mp34FromJfd, Mp34FromJff, Mp34FromJfo, Mp34FromJfq, Mp34FromJft, Mp34FromPmp, Mp34FromImacid)


def JfdFrom(Jfdfrom_ExcluMp34, JfdFromJff, JfdFromJfo, JfdFromJfq, JfdFromJft, JfdFromPmp, JfdFromImacid, JfdFromMp34):
    for i in AllEntiteJfd:
        Jfdfrom_.append(Jorf["JFD"].getEchanges(i))
    JfdFromJff.append(Jfdfrom_[0])
    JfdFromJfo.append(Jfdfrom_[1])
    JfdFromJfq.append(Jfdfrom_[2])
    JfdFromJft.append(Jfdfrom_[3])
    JfdFromPmp.append(Jfdfrom_[4])
    JfdFromImacid.append(Jfdfrom_[5])
    JfdFromMp34.append(Jfdfrom_[6])
    return (JfdFromJff, JfdFromJfo, JfdFromJfq, JfdFromJft, JfdFromPmp, JfdFromImacid, JfdFromMp34)


def JffFrom(jffFrom_, jffFromImacid, jffFromPmp, jffFromJft, jffFromJfq, jffFromJfo, jffFromJfd, jffFromMp34):
    for i in AllEntiteJff:
        jffFrom_.append(Jorf["JFF"].getEchanges(i))
    jffFromMp34.append(jffFrom_[0])
    jffFromJfd.append(jffFrom_[1])
    jffFromJfo.append(jffFrom_[2])
    jffFromJfq.append(jffFrom_[3])
    jffFromJft.append(jffFrom_[4])
    jffFromPmp.append(jffFrom_[5])
    jffFromImacid.append(jffFrom_[6])
    return (jffFromJfo, jffFromJfq, jffFromJft, JfdFromJft, JfdFromPmp, JfdFromImacid, jffFromMp34, jffFromJfd)


def insertionEchange(liste, conteur, nlingne, ncolone):
    conteur = 0
    for i in range(0, len(liste[0])):
        for k in range(0, len(liste[0][i])):
            for j in range(0, 4):
                Sheet1.cell(row=nlingne + k, column=ncolone + j + conteur).value = liste[0][i][k][j]
        conteur = conteur + 5


def JfoFrom(jfoFrom_, jfoFromImacid, jfoFromPmp, jfoFromJft, jfoFromJfq, jfoFromJff, jfoFromJfd, jfoFromMp34):
    for i in AllEntiteJfo:
        jfoFrom_.append(Jorf["JFO"].getEchanges(i))
    jfoFromMp34.append(jfoFrom_[0])
    jfoFromJfd.append(jfoFrom_[1])
    jfoFromJff.append(jfoFrom_[2])
    jfoFromJfq.append(jfoFrom_[3])
    jfoFromJft.append(jfoFrom_[4])
    jfoFromPmp.append(jfoFrom_[5])
    jfoFromImacid.append(jfoFrom_[6])
    return (jfoFromImacid, jfoFromPmp, jfoFromJft, jfoFromJfq, jfoFromJff, jfoFromJfd, jfoFromMp34)


def JfqFrom(jfqFrom_, jfqFromImacid, jfqFromPmp, jfqFromJft, jfqFromJfo, jfqFromJff, jfqFromJfd, jfqFromMp34):
    for i in AllEntiteJfq:
        jfqFrom_.append(Jorf["JFQ"].getEchanges(i))
    jfqFromMp34.append(jfqFrom_[0])
    jfqFromJfd.append(jfqFrom_[1])
    jfqFromJff.append(jfqFrom_[2])
    jfqFromJfo.append(jfqFrom_[3])
    jfqFromJft.append(jfqFrom_[4])
    jfqFromPmp.append(jfqFrom_[5])
    jfqFromImacid.append(jfqFrom_[6])
    return (jfqFromImacid, jfqFromPmp, jfqFromJft, jfqFromJfo, jfqFromJff, jfqFromJfd, jfqFromMp34)


def JftFrom(jftFrom_, jftFromImacid, jftFromPmp, jftFromJfq, jftFromJfo, jftFromJff, jftFromJfd, jftFromMp34):
    for i in AllEntiteJft:
        jftFrom_.append(Jorf["JFT"].getEchanges(i))
    jftFromMp34.append(jfqFrom_[0])
    jftFromJfd.append(jfqFrom_[1])
    jftFromJff.append(jfqFrom_[2])
    jftFromJfo.append(jfqFrom_[3])
    jftFromJfq.append(jfqFrom_[4])
    jftFromPmp.append(jfqFrom_[5])
    jftFromImacid.append(jfqFrom_[6])
    return (jftFromImacid, jftFromPmp, jftFromJfq, jftFromJfo, jftFromJff, jftFromJfd, jftFromMp34)


def PmpFrom(PmpFrom_, PmpFromImacid, PmpFromJft, PmpFromJfq, PmpFromJfo, PmpFromJff, PmpFromJfd, PmpFromMp34):
    for i in AllEntitePmp:
        PmpFrom_.append(Jorf["PMP"].getEchanges(i))
    PmpFromMp34.append(jfqFrom_[0])
    PmpFromJfd.append(jfqFrom_[1])
    PmpFromJff.append(jfqFrom_[2])
    PmpFromJfo.append(jfqFrom_[3])
    PmpFromJfq.append(jfqFrom_[4])
    PmpFromJft.append(jfqFrom_[5])
    PmpFromImacid.append(jfqFrom_[6])
    return (PmpFromImacid, PmpFromJft, PmpFromJfq, PmpFromJfo, PmpFromJff, PmpFromJfd, PmpFromMp34)


def imacidFrom(imacidFrom_, imacidFromPmp, imacidFromJft, imacidFromJfq, imacidFromJfo, imacidFromJff, imacidFromJfd,
               imacidFromMp34):
    for i in AllEntiteImacid:
        imacidFrom_.append(Jorf["IMACID"].getEchanges(i))
    imacidFromMp34.append(imacidFrom_[0])
    imacidFromJfd.append(imacidFrom_[1])
    imacidFromJff.append(imacidFrom_[2])
    imacidFromJfo.append(imacidFrom_[3])
    imacidFromJfq.append(imacidFrom_[4])
    imacidFromJft.append(imacidFrom_[5])
    imacidFromPmp.append(jfqFrom_[6])
    return (imacidFromPmp, imacidFromJft, imacidFromJfq, imacidFromJfo, imacidFromJff, imacidFromJfd, imacidFromMp34)


def inserRegime(listeRegime):
    for i in Jorf.getAllLignesDeProduction_():
        listeRegime.append(i.nom)

    listeRegime.insert(0, "0")

    listeRegime.insert(1, "0")

    listeRegime.insert(2, "0")

    listeRegime.insert(3, "0")

    listeRegime.insert(16, "0")

    listeRegime.insert(17, "0")

    listeRegime.insert(18, "0")

    return (listeRegime)


def maintenanceDisponibilite(maintenanceD0):
    date_time_obj = datetime.datetime.strptime(Jorf.date, '%d/%m/%Y')

    date_time = date_time_obj.strftime("%d-%m-%Y %H:%M:%S")

    end_date = date_time_obj + datetime.timedelta(days=Jorf.durée)

    end_datef = end_date.strftime("%d-%m-%Y %H:%M:%S")

    for ligneEngrais in Jorf['MP34'].lignesEngrais.values():

        d = ligneEngrais.disponibilité

        if d == 0:
            maintenanceD0.append([ligneEngrais.mappingMaintenance, date_time, end_datef, 0])

    for i in engraisJfExtraction:

        for ligneEngrais in Jorf[i].lignesEngrais.values():

            d = ligneEngrais.disponibilité

            if d == 0:
                maintenanceD0.append([ligneEngrais.mappingMaintenance, date_time, end_datef, 0])

    for atelier in Jorf['MP34'].ateliers.values():

        for ligne in atelier.lignesDeProduction.values():

            if ligne.disponibilité == 0:
                maintenanceD0.append([ligne.mappingMaintenance, date_time, end_datef, 0])

    for nomAtelier in Jorf.ENCHAINEMENT_ATELIER:

        for entité in Jorf.entités.values():

            if entité.nom == 'MP34':
                continue

            atelier = entité.getAtelier(nomAtelier)

            for ligne in atelier.lignesDeProduction.values():

                if ligne.disponibilité == 0:
                    maintenanceD0.append([ligne.mappingMaintenance, date_time, end_datef, 0])

    return (maintenanceD0)


# traitement
Sheet1["D3"] = Jorf.date
Sheet1["E1"] = Jorf.durée

infoEngraisMp34(engraisMp34_)
for i in engraisMp34_:
    Sheet1.cell(row=105 + conteur, column=2).value = i[0]
    Sheet1.cell(row=105 + conteur, column=3).value = i[1]
    conteur = conteur + 1

conteur = 0
infoEngraisJf(engraisJf_)
for i in engraisJf_:
    Sheet1.cell(row=105 + conteur, column=9).value = i[0]
    Sheet1.cell(row=105 + conteur, column=10).value = i[1]
    conteur = conteur + 1

carnetMp34(commandes, CommandesEngraisMp34_, conteur1)
for k in range(0, len(CommandesEngraisMp34_)):
    Sheet1.cell(row=134 + k, column=1).value = CommandesEngraisMp34_[k][0]
    Sheet1.cell(row=134 + k, column=2).value = CommandesEngraisMp34_[k][3]
    Sheet1.cell(row=134 + k, column=3).value = CommandesEngraisMp34_[k][1]
    Sheet1.cell(row=134 + k, column=4).value = CommandesEngraisMp34_[k][2]

carnetJf(commandesJf, conteur2, CommandesEngraisJf_)
for k in range(0, len(CommandesEngraisJf_)):
    Sheet1.cell(row=134 + k, column=8).value = CommandesEngraisJf_[k][0]
    Sheet1.cell(row=134 + k, column=9).value = CommandesEngraisJf_[k][3]
    Sheet1.cell(row=134 + k, column=10).value = CommandesEngraisJf_[k][1]
    Sheet1.cell(row=134 + k, column=11).value = CommandesEngraisJf_[k][2]

productionMp34(productionSulfurique, productionPhosphorique29, productionPhosphorique54, conteur3)
for i in range(0, len(productionSulfurique)):
    Sheet1.cell(row=222 + i, column=11).value = productionSulfurique[i][0]
    Sheet1.cell(row=222 + i, column=12).value = productionSulfurique[i][1]
for i in range(0, len(productionPhosphorique29)):
    Sheet1.cell(row=251 + i, column=11).value = productionPhosphorique29[i][0]
    Sheet1.cell(row=251 + i, column=12).value = productionPhosphorique29[i][1]
for i in range(0, len(productionPhosphorique54)):
    Sheet1.cell(row=281 + i, column=8).value = productionPhosphorique54[i][0]
    Sheet1.cell(row=281 + i, column=9).value = productionPhosphorique54[i][1]

conteur = 0
productionJf(productionSulfuriqueJf, productionPhosphorique29Jf, productionPhosphorique54Jf, conteur4)
for i in range(0, len(productionSulfuriqueJf)):
    conteur = conteur + 2
    Sheet1.cell(row=233 + conteur, column=11).value = productionSulfuriqueJf[i][0]
    Sheet1.cell(row=233 + conteur, column=12).value = productionSulfuriqueJf[i][1]
conteur = 0
for i in range(0, len(productionPhosphorique29Jf)):
    conteur = conteur + 2
    Sheet1.cell(row=260 + conteur, column=11).value = productionPhosphorique29Jf[i][0]
    Sheet1.cell(row=260 + conteur, column=12).value = productionPhosphorique29Jf[i][1]
conteur = 0
for i in range(0, len(productionPhosphorique54Jf)):
    Sheet1.cell(row=306 + conteur, column=8).value = productionPhosphorique54Jf[i][0]
    Sheet1.cell(row=306 + conteur, column=9).value = productionPhosphorique54Jf[i][1]
    conteur = conteur + 1
    Saut = [3, 8, 13, 18, 23, 28]
    if conteur in Saut:
        conteur = conteur + 2

capaciteUtilise(capaciteEntite, capaciteCommune)
for i in range(0, len(capaciteEntite)):
    Sheet1["F" + str(352 + i)] = capaciteEntite[i]
for i in range(0, len(capaciteCommune)):
    Sheet1["F" + str(433 + i)] = capaciteCommune[i]

# insertion souffre
echangeSouffre(AllSouffre, SouffreMp34, SouffreJfd, SouffreJff, SouffreJfo, SouffreJfq, SouffreJft, SouffrePmp,
               SouffreImacid)
for i in range(0, len(SouffreMp34[0])):
    for j in range(0, 4):
        Sheet1.cell(row=448 + i, column=2 + j).value = SouffreMp34[0][i][j]
for i in range(0, len(SouffreJfd[0])):
    for j in range(0, 4):
        Sheet1.cell(row=501 + i, column=2 + j).value = SouffreJfd[0][i][j]
for i in range(0, len(SouffreJff[0])):
    for j in range(0, 4):
        Sheet1.cell(row=448 + i, column=28 + j).value = SouffreJff[0][i][j]
for i in range(0, len(SouffreJfo[0])):
    for j in range(0, 4):
        Sheet1.cell(row=448 + i, column=49 + j).value = SouffreJfo[0][i][j]
for i in range(0, len(SouffreJfq[0])):
    for j in range(0, 4):
        Sheet1.cell(row=448 + i, column=70 + j).value = SouffreJfq[0][i][j]
for i in range(0, len(SouffreJft[0])):
    for j in range(0, 4):
        Sheet1.cell(row=448 + i, column=91 + j).value = SouffreJft[0][i][j]
for i in range(0, len(SouffrePmp[0])):
    for j in range(0, 4):
        Sheet1.cell(row=448 + i, column=112 + j).value = SouffrePmp[0][i][j]
for i in range(0, len(SouffreImacid[0])):
    for j in range(0, 4):
        Sheet1.cell(row=448 + i, column=133 + j).value = SouffreImacid[0][i][j]
    # insertion acide

# insertion importationMP34
Mp34From(Mp34from_, Mp34FromJfd, Mp34FromJff, Mp34FromJfo, Mp34FromJfq, Mp34FromJft, Mp34FromPmp, Mp34FromImacid)
insertionEchange(Mp34FromJfd, conteur, 448, 12)
insertionEchange(Mp34FromJff, conteur, 448, 33)
insertionEchange(Mp34FromJfo, conteur, 448, 54)
insertionEchange(Mp34FromJfq, conteur, 448, 75)
insertionEchange(Mp34FromJft, conteur, 448, 96)
insertionEchange(Mp34FromPmp, conteur, 448, 117)
insertionEchange(Mp34FromImacid, conteur, 448, 138)

# insertion importationJFD
JfdFrom(Jfdfrom_, JfdFromJff, JfdFromJfo, JfdFromJfq, JfdFromJft, JfdFromPmp, JfdFromImacid, JfdFromMp34)
insertionEchange(JfdFromImacid, conteur, 448, 154)
insertionEchange(JfdFromPmp, conteur, 448, 170)
insertionEchange(JfdFromJff, conteur, 448, 186)
insertionEchange(JfdFromJfo, conteur, 448, 202)
insertionEchange(JfdFromJfq, conteur, 448, 218)
insertionEchange(JfdFromJft, conteur, 448, 234)
insertionEchange(JfdFromMp34, conteur, 501, 7)

# insertion importationJFF
JffFrom(jffFrom_, jffFromImacid, jffFromPmp, jffFromJft, jffFromJfq, jffFromJfo, jffFromJfd, jffFromMp34)
insertionEchange(jffFromMp34, conteur, 501, 33)
insertionEchange(jffFromJfd, conteur, 501, 186)
insertionEchange(jffFromImacid, conteur, 448, 250)
insertionEchange(jffFromPmp, conteur, 448, 266)
insertionEchange(jffFromJfo, conteur, 448, 282)
insertionEchange(jffFromJfq, conteur, 448, 298)
insertionEchange(jffFromJft, conteur, 448, 314)

# insertion importationJFO
JfoFrom(jfoFrom_, jfoFromImacid, jfoFromPmp, jfoFromJft, jfoFromJfq, jfoFromJff, jfoFromJfd, jfoFromMp34)
insertionEchange(jfoFromMp34, conteur, 501, 54)
insertionEchange(jfoFromJfd, conteur, 501, 202)
insertionEchange(jfoFromJff, conteur, 501, 282)
insertionEchange(jfoFromImacid, conteur, 448, 330)
insertionEchange(jfoFromPmp, conteur, 448, 346)
insertionEchange(jfoFromJfq, conteur, 448, 362)
insertionEchange(jfoFromJft, conteur, 448, 378)

# insertion importantionJFQ
JfqFrom(jfqFrom_, jfqFromImacid, jfqFromPmp, jfqFromJft, jfqFromJfo, jfqFromJff, jfqFromJfd, jfqFromMp34)
insertionEchange(jfqFromMp34, conteur, 501, 75)
insertionEchange(jfqFromJfd, conteur, 501, 218)
insertionEchange(jfqFromJff, conteur, 501, 298)
insertionEchange(jfqFromJfo, conteur, 501, 362)
insertionEchange(jfqFromImacid, conteur, 448, 394)
insertionEchange(jfqFromPmp, conteur, 448, 410)
insertionEchange(jfqFromJft, conteur, 448, 426)

# insertion importationJFT
JftFrom(jftFrom_, jftFromImacid, jftFromPmp, jftFromJff, jftFromJfo, jftFromJfq, jftFromJfd, jftFromMp34)
insertionEchange(jftFromMp34, conteur, 501, 96)
insertionEchange(jftFromJfd, conteur, 501, 234)
insertionEchange(jftFromJff, conteur, 501, 314)
insertionEchange(jftFromJfo, conteur, 501, 378)
insertionEchange(jftFromJfq, conteur, 501, 426)
insertionEchange(jftFromImacid, conteur, 448, 442)
insertionEchange(jftFromPmp, conteur, 448, 458)

# insertion importationPMP
PmpFrom(PmpFrom_, PmpFromImacid, PmpFromJft, PmpFromJfq, PmpFromJfo, PmpFromJff, PmpFromJfd, PmpFromMp34)
insertionEchange(PmpFromMp34, conteur, 501, 117)
insertionEchange(PmpFromJfd, conteur, 501, 170)
insertionEchange(PmpFromJff, conteur, 501, 266)
insertionEchange(PmpFromJfo, conteur, 501, 346)
insertionEchange(PmpFromJfq, conteur, 501, 410)
insertionEchange(PmpFromJft, conteur, 501, 458)
insertionEchange(PmpFromImacid, conteur, 501, 474)

# insertion importationIMACID
imacidFrom(imacidFrom_, imacidFromPmp, imacidFromJft, imacidFromJfq, imacidFromJfo, imacidFromJff, imacidFromJfd,
           imacidFromMp34)
insertionEchange(imacidFromMp34, conteur, 501, 138)
insertionEchange(imacidFromJfd, conteur, 501, 154)
insertionEchange(imacidFromJff, conteur, 501, 250)
insertionEchange(imacidFromJfo, conteur, 501, 330)
insertionEchange(imacidFromJfq, conteur, 501, 394)
insertionEchange(imacidFromJft, conteur, 501, 442)
insertionEchange(imacidFromPmp, conteur, 448, 474)

# insertion externe
sulfuriqueExterne = Jorf["MP34"].getEchanges('extérieur')[1]
for i in range(0, len(sulfuriqueExterne)):
    for j in range(0, 4):
        Sheet1.cell(row=448 + i, column=7 + j).value = sulfuriqueExterne[i][j]
phosphorique54Externe = Jorf.getExportMP34()
for i in range(0, len(phosphorique54Externe)):
    for j in range(0, 4):
        Sheet1.cell(row=501 + i, column=22 + j).value = phosphorique54Externe[i][j]

# insertion maintenances
Sheet2 = wb["Scheduled Maintenance1"]
insertionMaintenance = Jorf.getMaintenances()
for i in range(0, len(insertionMaintenance)):
    for j in range(0, 4):
        Sheet2.cell(row=1 + i, column=1 + j).value = insertionMaintenance[i][j]




#insertion des ligne non disponnible

maintenanceDisponibilite(maintenanceD0)

for i in range (0,len(maintenanceD0)):

    for j in range(0,4):

        Sheet2.cell(row=1+len(insertionMaintenance)+i,column=1+j).value=maintenanceD0[i][j]



#insertion regimeDeMarche

Sheet3 = wb["Regime de marche"]

inserRegime(listeRegime)

conteur=0

for r in Jorf.emploiRégimeDeMarche:

    Sheet3.cell(row=2,column=3+conteur).value=r

    for i in Jorf.emploiRégimeDeMarche[r]:

        Sheet3.cell(row=listeRegime.index(i[0]),column=3+conteur).value=i[1]

    conteur=conteur+1


wb.save("C:/Users/HP/Desktop/test3.xlsm")